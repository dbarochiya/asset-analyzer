#!/usr/bin/env python3
"""
parse_data.py — reads raw-data.xlsx (5 raw fields only), derives all metrics,
                computes per-row MF-style NAV and monthly XIRR, writes data.json
"""

import json
import openpyxl
from datetime import datetime, timedelta
import pandas as pd
import yfinance as yf

XLSX_PATH = "raw-data.xlsx"
OUTPUT_PATH = "data.json"

# ── helpers ───────────────────────────────────────────────────────────────────

def to_epoch_ms(dt):
    return int(dt.timestamp() * 1000) if isinstance(dt, datetime) else None

def r2(v):
    return round(v, 2) if v is not None else None

def r4(v):
    return round(v, 4) if v is not None else None

# ── Newton-Raphson XIRR ───────────────────────────────────────────────────────

def xirr(cashflows, dates, guess=0.1, max_iter=1000, tol=1e-6):
    """
    cashflows: list of floats  (positive = inflow, negative = exit / outflow)
    dates:     list of datetime objects, same length
    Returns annualised rate as a decimal, or None on failure.
    """
    if not cashflows or len(cashflows) < 2:
        return None
    # Day fractions relative to first date
    t0 = dates[0]
    days = [(d - t0).days / 365.0 for d in dates]

    rate = guess
    for _ in range(max_iter):
        npv = sum(cf / (1 + rate) ** t for cf, t in zip(cashflows, days))
        d_npv = sum(-t * cf / (1 + rate) ** (t + 1) for cf, t in zip(cashflows, days))
        if d_npv == 0:
            return None
        new_rate = rate - npv / d_npv
        if abs(new_rate - rate) < tol:
            return new_rate
        rate = new_rate
        # Guard against divergence
        if rate <= -1:
            rate = -0.999
    return None  # did not converge

# ── parse one sheet (raw fields only) ────────────────────────────────────────

RAW_COLS = {"date": 0, "cashflow": 2, "free_cash": 3, "invested": 4, "pv": 5}

def parse_sheet(ws):
    """
    Read only the 5 raw columns. Skip the XIRR-metadata row (last row
    where col B is a float, not a string platform name).
    Returns list of raw dicts.
    """
    rows = list(ws.iter_rows(values_only=True))
    # row[0] is header; skip it
    raw_rows = []
    for row in rows[1:]:
        dt = row[0]
        platform = row[1]
        # XIRR metadata row: platform cell is a float
        if isinstance(platform, float) or dt is None:
            continue
        # Skip rows with no meaningful data
        if row[2] is None and row[4] is None and row[5] is None:
            continue

        cf   = float(row[2]) if row[2] is not None else 0.0
        fc   = float(row[3]) if row[3] is not None else 0.0
        inv  = float(row[4]) if row[4] is not None else 0.0
        pv   = float(row[5]) if row[5] is not None else 0.0

        raw_rows.append({
            "date":      dt.strftime("%Y-%m-%d"),
            "epoch_ms":  to_epoch_ms(dt),
            "_dt":       dt,   # kept for XIRR date math, stripped before JSON
            "cashflow":  cf,
            "free_cash": fc,
            "invested":  inv,
            "pv":        pv,
        })
    return raw_rows

# ── derive all calculated fields ──────────────────────────────────────────────

def derive_fields(raw_rows):
    """
    Given raw rows (date, cashflow, free_cash, invested, pv),
    compute all derived fields INCLUDING MF-style NAV and monthly XIRR.
    Returns enriched row list.
    """
    rows = []
    cum_invested = 0.0

    # MF-style NAV state
    units = None  # initialised on first row

    for i, raw in enumerate(raw_rows):
        cf   = raw["cashflow"]
        fc   = raw["free_cash"]
        inv  = raw["invested"]
        pv   = raw["pv"]

        # ── Cumulative / derived ─────────────────────────────────────────
        cum_invested += cf
        total_pv          = pv + fc
        gain_unrealised   = pv - inv
        gain_realised     = inv + fc - cum_invested
        gain_total        = gain_unrealised + gain_realised
        returns_pct       = (gain_total / cum_invested * 100) if cum_invested else 0.0

        # ── MF-style NAV ─────────────────────────────────────────────────
        # Seed: on day 0 we issue units such that NAV = 100
        # Subsequently, each cashflow buys (or redeems) units at current NAV
        if units is None:
            # Initialise: total_pv / 100 units → NAV = 100
            units = total_pv / 100.0
            nav   = 100.0
        else:
            # New cashflow buys units at the previous NAV
            if nav_prev > 0:
                new_units = cf / nav_prev
            else:
                new_units = 0.0
            units += new_units
            nav = total_pv / units if units > 0 else 100.0

        nav_prev = nav

        # ── Monthly XIRR (cumulative to this date) ───────────────────────
        # Simulate: all cashflows[0..i] as inflows, exit at total_pv[i]
        cf_series   = [row["cashflow"] for row in raw_rows[:i+1]]
        date_series = [row["_dt"]      for row in raw_rows[:i+1]]
        # Add terminal outflow (liquidation at this PV)
        cf_series.append(-total_pv)
        date_series.append(raw["_dt"])

        if i == 0:
            # Single-period: simple return
            monthly_xirr_pct = returns_pct
        else:
            xirr_val = xirr(cf_series, date_series)
            monthly_xirr_pct = round(xirr_val * 100, 4) if xirr_val is not None else None

        rows.append({
            "date":              raw["date"],
            "epoch_ms":          raw["epoch_ms"],
            # — raw fields —
            "cashflow":          r2(cf),
            "free_cash":         r2(fc),
            "invested":          r2(inv),
            "pv":                r2(pv),
            # — derived fields —
            "total_invested":    r2(cum_invested),
            "total_pv":          r2(total_pv),
            "gain_unrealised":   r2(gain_unrealised),
            "gain_realised":     r2(gain_realised),
            "gain_total":        r2(gain_total),
            "returns_pct":       r4(returns_pct),
            "nav":               r4(nav),
            "monthly_xirr_pct":  r4(monthly_xirr_pct),
        })

    return rows

# ── account-level analytics ───────────────────────────────────────────────────

def compute_analytics(rows, name):
    if not rows:
        return {}

    pvs      = [r["total_pv"]    for r in rows]
    returns  = [r["returns_pct"] for r in rows]
    navs     = [r["nav"]         for r in rows]

    best_idx  = returns.index(max(returns))
    worst_idx = returns.index(min(returns))

    # Max drawdown on total_pv
    peak = pvs[0];  max_dd = 0.0
    for tv in pvs:
        if tv > peak: peak = tv
        dd = (peak - tv) / peak * 100 if peak else 0
        if dd > max_dd: max_dd = dd

    # Consecutive negative months
    max_neg = cur = 0
    for ret in returns:
        cur = cur + 1 if ret < 0 else 0
        max_neg = max(max_neg, cur)

    # Breakeven date (first non-negative after having been negative)
    breakeven = None
    was_neg = False
    for row in rows:
        if row["returns_pct"] < 0:
            was_neg = True
        elif was_neg and row["returns_pct"] >= 0:
            breakeven = row["date"]
            break

    # Overall XIRR (all cashflows + exit at last total_pv)
    cf_all   = [r["cashflow"] for r in rows] + [-rows[-1]["total_pv"]]
    dt_all   = [datetime.strptime(r["date"], "%Y-%m-%d") for r in rows] + \
               [datetime.strptime(rows[-1]["date"], "%Y-%m-%d")]
    xirr_val = xirr(cf_all, dt_all)
    xirr_pct = round(xirr_val * 100, 4) if xirr_val is not None else None

    avg_ret  = sum(returns) / len(returns)
    variance = sum((x - avg_ret) ** 2 for x in returns) / len(returns)
    vol      = variance ** 0.5

    last = rows[-1]
    return {
        "name":                          name,
        "xirr_pct":                      xirr_pct,
        "start_date":                    rows[0]["date"],
        "end_date":                      last["date"],
        "num_months":                    len(rows),
        "final_total_invested":          last["total_invested"],
        "final_pv":                      last["pv"],
        "final_total_pv":                last["total_pv"],
        "final_gain_total":              last["gain_total"],
        "final_gain_unrealised":         last["gain_unrealised"],
        "final_gain_realised":           last["gain_realised"],
        "final_returns_pct":             last["returns_pct"],
        "final_nav":                     last["nav"],
        "best_month":  {"date": rows[best_idx]["date"],  "returns_pct": returns[best_idx],  "gain_total": rows[best_idx]["gain_total"]},
        "worst_month": {"date": rows[worst_idx]["date"], "returns_pct": returns[worst_idx], "gain_total": rows[worst_idx]["gain_total"]},
        "max_drawdown_pct":              r2(max_dd),
        "max_consecutive_negative_months": max_neg,
        "breakeven_date":                breakeven,
        "avg_monthly_return_pct":        r2(avg_ret),
        "volatility_pct":                r2(vol),
    }

# ── combined portfolio (across accounts) ──────────────────────────────────────

def compute_combined(accounts_data):
    date_map = {}
    for acc in accounts_data:
        for row in acc["rows"]:
            d = row["date"]
            if d not in date_map:
                date_map[d] = {"date": d, "epoch_ms": row["epoch_ms"], "accounts": {}}
            date_map[d]["accounts"][acc["name"]] = row

    sorted_dates = sorted(date_map.keys())
    combined_rows = []
    last_known = {acc["name"]: None for acc in accounts_data}

    # Combined MF-style NAV state
    c_units = None
    c_nav_prev = 100.0

    for d in sorted_dates:
        entry = date_map[d]
        tot_invested = tot_pv = tot_gain = tot_cf = 0.0

        for acc in accounts_data:
            n = acc["name"]
            if n in entry["accounts"]:
                last_known[n] = entry["accounts"][n]
            if last_known[n]:
                rv = last_known[n]
                tot_invested += rv.get("total_invested") or 0
                tot_pv       += rv.get("total_pv") or 0
                tot_gain     += rv.get("gain_total") or 0
                tot_cf       += rv.get("cashflow") or 0

        returns_pct = r4(tot_gain / tot_invested * 100) if tot_invested else 0

        # Combined MF-style NAV
        if c_units is None:
            c_units    = tot_pv / 100.0
            c_nav      = 100.0
        else:
            if c_nav_prev > 0:
                c_units += tot_cf / c_nav_prev
            c_nav = tot_pv / c_units if c_units > 0 else 100.0
        c_nav_prev = c_nav

        combined_rows.append({
            "date":           d,
            "epoch_ms":       entry["epoch_ms"],
            "total_invested": r2(tot_invested),
            "total_pv":       r2(tot_pv),
            "total_gain":     r2(tot_gain),
            "cashflow":       r2(tot_cf),
            "returns_pct":    returns_pct,
            "nav":            r4(c_nav),
        })

    final = combined_rows[-1]
    return {
        "rows": combined_rows,
        "final_total_invested": final["total_invested"],
        "final_total_pv":       final["total_pv"],
        "final_total_gain":     final["total_gain"],
        "final_returns_pct":    final["returns_pct"],
        "final_nav":            final["nav"],
        "total_months_tracked": len(combined_rows),
    }

# ── insights ──────────────────────────────────────────────────────────────────

def generate_insights(analytics_list, combined):
    ins = []
    for a in analytics_list:
        ins.append(f"📈 Best month for {a['name']}: {a['best_month']['date']} with {a['best_month']['returns_pct']:+.2f}% return (₹{a['best_month']['gain_total']:,.0f} gain)")
        ins.append(f"📉 Worst month for {a['name']}: {a['worst_month']['date']} with {a['worst_month']['returns_pct']:+.2f}% return")
        if a['max_consecutive_negative_months'] > 0:
            ins.append(f"⚠️ {a['name']} was in negative territory for {a['max_consecutive_negative_months']} consecutive month(s)")
        if a['breakeven_date']:
            ins.append(f"✅ {a['name']} recovered to breakeven by {a['breakeven_date']}")
        ins.append(f"📊 {a['name']} XIRR: {a['xirr_pct']:+.2f}% annualised")
        ins.append(f"🎯 {a['name']} max drawdown: -{a['max_drawdown_pct']:.2f}%")
        ins.append(f"📐 {a['name']} avg monthly return: {a['avg_monthly_return_pct']:+.2f}% | Volatility: {a['volatility_pct']:.2f}%")
        ins.append(f"📌 {a['name']} final NAV: {a['final_nav']:.2f} (started at 100.00)")

    best_xirr = max(analytics_list, key=lambda a: a['xirr_pct'] or -999)
    ins.append(f"🏆 {best_xirr['name']} has the highest XIRR at {best_xirr['xirr_pct']:+.2f}%")
    ins.append(f"💼 Combined: ₹{combined['final_total_invested']:,.0f} invested → ₹{combined['final_total_pv']:,.0f} value "
               f"(₹{combined['final_total_gain']:,.0f} gain, {combined['final_returns_pct']:+.2f}%)")
    return ins

# ── benchmarking ──────────────────────────────────────────────────────────────

def fetch_benchmarks(combined_rows):
    benchmarks = {
        "Nifty 50": "^NSEI",
        "Nasdaq": "^IXIC",
        "Gold": "GC=F",
        "Smallcap 100": "^CNXSC",
        "Midcap 150": "^NSEMDCP50"
    }
    
    start_date_str = combined_rows[0]['date']
    end_date_str = combined_rows[-1]['date']
    
    start_dt = datetime.strptime(start_date_str, "%Y-%m-%d")
    end_dt = datetime.strptime(end_date_str, "%Y-%m-%d")
    
    # fetch a bit earlier/later to ensure we capture the bounds
    fetch_start = (start_dt - timedelta(days=32)).strftime("%Y-%m-01")
    fetch_end = (end_dt + timedelta(days=32)).strftime("%Y-%m-%d")
    
    tickers = " ".join(benchmarks.values())
    print("\nFetching global benchmarks via yfinance...")
    try:
        # download all
        df = yf.download(tickers, start=fetch_start, end=fetch_end, interval="1mo", progress=False)['Close']
    except Exception as e:
        print(f"Failed to fetch benchmarks: {e}")
        return []

    results = []
    
    for name, ticker in benchmarks.items():
        try:
            if ticker not in df.columns:
                series = df.dropna() if len(benchmarks) == 1 else df
            else:
                series = df[ticker].dropna()
                
            if series.empty:
                continue
                
            s_data = series[series.index >= start_date_str]
            end_match = end_dt.strftime("%Y-%m-31")
            e_data = s_data[s_data.index <= end_match]
            
            if s_data.empty or e_data.empty:
                continue
                
            start_price = float(s_data.iloc[0])
            end_price = float(e_data.iloc[-1])
            
            final_nav = 100.0 * (end_price / start_price)
            
            days = (e_data.index[-1] - s_data.index[0]).days
            if days > 0:
                years = days / 365.25
                cagr = ((end_price / start_price) ** (1/years) - 1) * 100
            else:
                cagr = 0.0
                
            # Compute monthly NAV progression mapping: "YYYY-MM" -> Base 100
            prog_map = {}
            for dt, p in e_data.items():
                prog_map[dt.strftime("%Y-%m")] = round(100.0 * (float(p) / start_price), 4)
                
            results.append({
                "name": name,
                "final_nav": round(final_nav, 2),
                "cagr_pct": round(cagr, 2),
                "nav_progression": prog_map
            })
        except Exception as e:
            print(f"Skipping {name}: {e}")
            
    # Fixed Deposit Benchmark
    fd_prog_map = {}
    for r in combined_rows:
        dt_obj = datetime.strptime(r['date'], "%Y-%m-%d")
        days = (dt_obj - start_dt).days
        years_c = days / 365.25 if days > 0 else 0
        fd_nav_t = 100.0 * ((1 + 0.07)**years_c)
        fd_prog_map[r['date'][:7]] = round(fd_nav_t, 4)
        
    days_total = (end_dt - start_dt).days
    years_total = days_total / 365.25 if days_total > 0 else 0
    fd_final_nav = 100.0 * ((1 + 0.07)**years_total)
    
    results.append({
        "name": "Fixed Deposit (7%)",
        "final_nav": round(fd_final_nav, 2),
        "cagr_pct": 7.00,
        "nav_progression": fd_prog_map
    })
    
    return sorted(results, key=lambda x: x['final_nav'], reverse=True)

# ── main ──────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    accounts_data  = []
    analytics_list = []

    for sheet_name in wb.sheetnames:
        ws       = wb[sheet_name]
        raw_rows = parse_sheet(ws)
        rows     = derive_fields(raw_rows)
        analytics = compute_analytics(rows, sheet_name)
        accounts_data.append({"name": sheet_name, "rows": rows})
        analytics_list.append(analytics)
        print(f"✓ {sheet_name}: {len(rows)} rows | XIRR={analytics['xirr_pct']}% | final NAV={analytics['final_nav']}")

    combined = compute_combined(accounts_data)
    insights = generate_insights(analytics_list, combined)

    # Global benchmarks mapped exactly to combined start and end
    benchmarks = fetch_benchmarks(combined['rows'])

    output = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "accounts": [
            {"name": acc["name"], "rows": acc["rows"], "analytics": al}
            for acc, al in zip(accounts_data, analytics_list)
        ],
        "combined": combined,
        "insights": insights,
        "benchmarks": benchmarks,
    }

    with open(OUTPUT_PATH, "w") as f:
        json.dump(output, f, indent=2, default=str)

    print(f"\n✓ {OUTPUT_PATH} written")
    print(f"  Combined: ₹{combined['final_total_invested']:,.0f} → ₹{combined['final_total_pv']:,.0f} "
          f"(gain ₹{combined['final_total_gain']:,.0f}) | NAV {combined['final_nav']:.2f}")


if __name__ == "__main__":
    main()
